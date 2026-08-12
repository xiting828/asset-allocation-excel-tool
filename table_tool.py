"""粘贴表格解析和 Excel 生成逻辑。"""

from __future__ import annotations

import csv
import math
import re
import unicodedata
from dataclasses import dataclass
from datetime import datetime
from io import BytesIO, StringIO
from typing import Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


INTEGER_RE = re.compile(r"^[+-]?(?:0|[1-9]\d*)$")
DECIMAL_RE = re.compile(r"^[+-]?(?:0|[1-9]\d*)\.\d+$")
MAX_INPUT_BYTES = 2 * 1024 * 1024
MAX_ROWS = 10_000
MAX_COLUMNS = 200
MAX_CELLS = 500_000
MAX_CELL_CHARACTERS = 32_000

# 只对明确指定的两个资产大类填色，颜色保持低饱和、便于阅读。
ASSET_CATEGORY_COLORS: dict[str, str] = {
    "固定收益类": "E8F3EC",  # 淡绿色
    "权益类": "DDEBF7",  # 淡蓝色
}


@dataclass(frozen=True)
class ParseResult:
    """解析后的矩形表格及相关提示信息。"""

    rows: list[list[str]]
    inconsistent_columns: bool
    original_column_counts: tuple[int, ...]

    @property
    def is_empty(self) -> bool:
        return not self.rows

    @property
    def column_count(self) -> int:
        return len(self.rows[0]) if self.rows else 0


class TableLimitError(ValueError):
    """输入表格超过安全处理限制。"""


def _is_blank(value: str) -> bool:
    return value.strip() == ""


def parse_pasted_table(text: str) -> ParseResult:
    """解析 Tab/换行分隔文本，替换星号，并只裁剪外围空白。"""

    if not text or not text.strip():
        return ParseResult([], False, ())

    if len(text.encode("utf-8")) > MAX_INPUT_BYTES:
        raise TableLimitError("粘贴内容不能超过 2 MB，请拆分后分批处理。")

    normalized = text.replace("\r\n", "\n").replace("\r", "\n")
    reader = csv.reader(StringIO(normalized), delimiter="\t", quoting=csv.QUOTE_MINIMAL)
    raw_rows: list[list[str]] = []
    total_cells = 0
    for row_number, row in enumerate(reader, start=1):
        if row_number > MAX_ROWS:
            raise TableLimitError(f"表格不能超过 {MAX_ROWS:,} 行，请拆分后分批处理。")
        if len(row) > MAX_COLUMNS:
            raise TableLimitError(f"表格不能超过 {MAX_COLUMNS} 列，请减少列数后重试。")
        total_cells += len(row)
        if total_cells > MAX_CELLS:
            raise TableLimitError(f"表格不能超过 {MAX_CELLS:,} 个单元格，请拆分后分批处理。")
        if any(len(cell) > MAX_CELL_CHARACTERS for cell in row):
            raise TableLimitError("单个单元格内容不能超过 32,000 个字符，请缩短后重试。")
        raw_rows.append([cell.replace("*", "0") for cell in row])

    # csv.reader 会为结尾换行产生恰当的行；这里只移除外围全空白行。
    while raw_rows and all(_is_blank(cell) for cell in raw_rows[0]):
        raw_rows.pop(0)
    while raw_rows and all(_is_blank(cell) for cell in raw_rows[-1]):
        raw_rows.pop()

    if not raw_rows:
        return ParseResult([], False, ())

    counts = tuple(len(row) for row in raw_rows)
    max_columns = max(counts)
    rows = [row + [""] * (max_columns - len(row)) for row in raw_rows]

    # 只裁剪所有行都为空白的最左/最右列，内部空行和空单元格原样保留。
    left = 0
    while left < max_columns and all(_is_blank(row[left]) for row in rows):
        left += 1
    right = max_columns
    while right > left and all(_is_blank(row[right - 1]) for row in rows):
        right -= 1

    rows = [row[left:right] for row in rows]
    if not rows or not rows[0]:
        return ParseResult([], len(set(counts)) > 1, counts)

    return ParseResult(rows, len(set(counts)) > 1, counts)


def display_width(value: object) -> int:
    """估算文本显示宽度；中日韩全角字符按两个英文字符计算。"""

    text = "" if value is None else str(value)
    widths = []
    for line in text.splitlines() or [""]:
        widths.append(
            sum(2 if unicodedata.east_asian_width(char) in {"W", "F"} else 1 for char in line)
        )
    return max(widths, default=0)


def safe_excel_value(value: str) -> str | int | float:
    """只转换明确且不会损失精度的数字，其余内容优先保留为文本。"""

    stripped = value.strip()
    if stripped == "":
        return ""

    signless = stripped.lstrip("+-")
    if INTEGER_RE.fullmatch(stripped):
        # 有前导零、超出 Excel 可靠精度的长编号均保留为文本。
        if (len(signless) > 1 and signless.startswith("0")) or len(signless) > 15:
            return value
        return int(stripped)

    if DECIMAL_RE.fullmatch(stripped):
        integer_part = signless.split(".", 1)[0]
        significant_digits = len(signless.replace(".", "").lstrip("0"))
        if (len(integer_part) > 1 and integer_part.startswith("0")) or significant_digits > 15:
            return value
        return float(stripped)

    return value


def _make_unique_headers(headers: Sequence[str]) -> list[str]:
    """为网页预览生成非空且不重复的列名，不改变导出的原始表头。"""

    used: dict[str, int] = {}
    result: list[str] = []
    for index, header in enumerate(headers, start=1):
        base = header.strip() or f"列{index}"
        used[base] = used.get(base, 0) + 1
        result.append(base if used[base] == 1 else f"{base}_{used[base]}")
    return result


def preview_parts(rows: Sequence[Sequence[str]], first_row_is_header: bool) -> tuple[list[str], list[list[str]]]:
    """返回网页预览使用的列名和数据行。"""

    if not rows:
        return [], []
    column_count = len(rows[0])
    if first_row_is_header:
        return _make_unique_headers(rows[0]), [list(row) for row in rows[1:]]
    return [f"列{index}" for index in range(1, column_count + 1)], [list(row) for row in rows]


def _estimate_wrapped_lines(value: object, usable_width: int) -> int:
    text = "" if value is None else str(value)
    lines = 0
    for part in text.splitlines() or [""]:
        part_width = max(display_width(part), 1)
        lines += max(1, math.ceil(part_width / max(usable_width, 1)))
    return lines


def find_asset_category_column(
    rows: Sequence[Sequence[object]], first_row_is_header: bool = True
) -> int | None:
    """返回“资产大类”列的零基索引；没有明确表头时返回 None。"""

    if not first_row_is_header or not rows:
        return None
    for index, header in enumerate(rows[0]):
        if str(header).strip() == "资产大类":
            return index
    return None


def asset_category_color(value: object) -> str | None:
    """仅在去除前后空格后精确匹配指定资产大类。"""

    return ASSET_CATEGORY_COLORS.get(str(value).strip())


def build_excel(rows: Sequence[Sequence[str]], first_row_is_header: bool = True) -> BytesIO:
    """在内存中生成经过排版的 xlsx 文件。"""

    if not rows or not rows[0]:
        raise ValueError("没有可导出的表格数据")

    column_count = len(rows[0])
    if first_row_is_header:
        export_rows = [list(row) for row in rows]
    else:
        export_rows = [
            [f"列{index}" for index in range(1, column_count + 1)],
            *[list(row) for row in rows],
        ]

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "整理后数据"
    sheet.freeze_panes = "A2"
    sheet.sheet_view.showGridLines = False

    for row_index, source_row in enumerate(export_rows, start=1):
        for column_index in range(1, column_count + 1):
            source = source_row[column_index - 1] if column_index <= len(source_row) else ""
            value: str | int | float = source if row_index == 1 else safe_excel_value(source)
            cell = sheet.cell(row=row_index, column=column_index, value=value)
            # 以等号开头的粘贴文本按文本保存，避免意外执行为 Excel 公式。
            if isinstance(value, str):
                cell.data_type = "s"
                cell.number_format = "@"

    max_row = len(export_rows)
    max_column = column_count
    asset_category_column = find_asset_category_column(rows, first_row_is_header)

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    # 等线简洁、稳重，适合中文商务表格；不指定字体颜色，保留 Excel 默认黑色。
    header_font = Font(name="等线", size=11, bold=True)
    body_font = Font(name="等线", size=10, bold=False)
    thin = Side(style="thin", color="B8C4CC")
    medium = Side(style="medium", color="7EA9C4")

    column_widths: list[float] = []
    for column_index in range(1, max_column + 1):
        widest = max(display_width(sheet.cell(row=row, column=column_index).value) for row in range(1, max_row + 1))
        width = min(35, max(8, widest + 2))
        column_widths.append(float(width))
        sheet.column_dimensions[get_column_letter(column_index)].width = width

    for row_index in range(1, max_row + 1):
        max_lines = 1
        category_color = (
            asset_category_color(export_rows[row_index - 1][asset_category_column])
            if (
                row_index > 1
                and asset_category_column is not None
                and asset_category_column < len(export_rows[row_index - 1])
            )
            else None
        )
        category_fill = PatternFill("solid", fgColor=category_color) if category_color else None
        for column_index in range(1, max_column + 1):
            cell = sheet.cell(row=row_index, column=column_index)
            is_header = row_index == 1
            needs_wrap = "\n" in str(cell.value or "") or display_width(cell.value) > max(column_widths[column_index - 1] - 2, 1)
            if is_header:
                cell.fill = header_fill
                cell.font = header_font
            else:
                cell.font = body_font
                if category_fill:
                    cell.fill = category_fill

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=needs_wrap,
            )
            cell.border = Border(
                left=medium if column_index == 1 else thin,
                right=medium if column_index == max_column else thin,
                top=medium if row_index == 1 else thin,
                bottom=medium if row_index == max_row else thin,
            )
            if not is_header:
                max_lines = max(
                    max_lines,
                    _estimate_wrapped_lines(cell.value, int(max(column_widths[column_index - 1] - 2, 1))),
                )

        sheet.row_dimensions[row_index].height = 26 if row_index == 1 else min(100, max(20, 18 * max_lines))

    sheet.auto_filter.ref = f"A1:{get_column_letter(max_column)}{max_row}"
    sheet.print_title_rows = "1:1"

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def download_filename(now: datetime | None = None) -> str:
    """生成包含当前日期和时间的下载文件名。"""

    current = now or datetime.now()
    return f"整理后数据_{current:%Y%m%d_%H%M%S}.xlsx"
