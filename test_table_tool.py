from __future__ import annotations

import unittest
from datetime import datetime

from openpyxl import load_workbook

from table_tool import (
    asset_category_color,
    build_excel,
    display_width,
    download_filename,
    find_asset_category_column,
    parse_pasted_table,
    preview_parts,
    safe_excel_value,
    TableLimitError,
)


SAMPLE = (
    "\t\t\t\t\n"
    "\t姓名\t学号\t成绩\t备注\t\n"
    "\t张三\t00123\t98\t这是用于验证自动换行和行高调整的很长中文文本，内容需要保持完整且不能乱码。\t\n"
    "\t李四\t00007\t*\t\t\n"
    "\t王五\t00999\t***\n"
    "\t\t\t\t\t\n"
    "\t赵六\t1234567890123456\t88.5\t普通文字\t\n"
    "\t\t\t\t\t\n"
)


class ParseTableTests(unittest.TestCase):
    def test_parse_mixed_data_and_only_trim_outer_blank(self) -> None:
        result = parse_pasted_table(SAMPLE)

        self.assertFalse(result.is_empty)
        self.assertTrue(result.inconsistent_columns)
        self.assertEqual(result.column_count, 4)
        self.assertEqual(result.rows[0], ["姓名", "学号", "成绩", "备注"])
        self.assertEqual(result.rows[2], ["李四", "00007", "0", ""])
        self.assertEqual(result.rows[3], ["王五", "00999", "000", ""])
        self.assertEqual(result.rows[4], ["", "", "", ""])
        self.assertEqual(result.rows[5][1], "1234567890123456")

    def test_empty_input(self) -> None:
        self.assertTrue(parse_pasted_table("  \n\t \n").is_empty)

    def test_preview_without_header_preserves_first_row_as_data(self) -> None:
        rows = [["甲", "乙"], ["1", "2"]]
        headers, data = preview_parts(rows, first_row_is_header=False)
        self.assertEqual(headers, ["列1", "列2"])
        self.assertEqual(data[0], ["甲", "乙"])

    def test_duplicate_or_blank_preview_headers_are_safe(self) -> None:
        headers, _ = preview_parts([["姓名", "姓名", ""], ["甲", "乙", "丙"]], True)
        self.assertEqual(headers, ["姓名", "姓名_2", "列3"])

    def test_rejects_too_many_rows(self) -> None:
        text = "表头\n" + "\n".join("数据" for _ in range(10_000))
        with self.assertRaisesRegex(TableLimitError, "10,000 行"):
            parse_pasted_table(text)

    def test_rejects_too_many_columns(self) -> None:
        with self.assertRaisesRegex(TableLimitError, "200 列"):
            parse_pasted_table("\t".join("值" for _ in range(201)))

    def test_rejects_oversized_cell(self) -> None:
        with self.assertRaisesRegex(TableLimitError, "32,000"):
            parse_pasted_table("字" * 32_001)


class ValueTests(unittest.TestCase):
    def test_safe_number_conversion(self) -> None:
        self.assertEqual(safe_excel_value("98"), 98)
        self.assertEqual(safe_excel_value("88.5"), 88.5)
        self.assertEqual(safe_excel_value("00123"), "00123")
        self.assertEqual(safe_excel_value("1234567890123456"), "1234567890123456")
        self.assertEqual(safe_excel_value("2026-08-12"), "2026-08-12")

    def test_chinese_display_width(self) -> None:
        self.assertEqual(display_width("中文A"), 5)

    def test_timestamped_filename(self) -> None:
        self.assertEqual(
            download_filename(datetime(2026, 8, 12, 9, 8, 7)),
            "整理后数据_20260812_090807.xlsx",
        )

    def test_asset_category_colors(self) -> None:
        self.assertEqual(asset_category_color(" 固定收益类 "), "E8F3EC")
        self.assertEqual(asset_category_color("权益类"), "DDEBF7")
        self.assertIsNone(asset_category_color("固定收益"))
        self.assertIsNone(asset_category_color("现金类"))

    def test_find_asset_category_column_is_exact(self) -> None:
        self.assertEqual(
            find_asset_category_column([["客户", " 资产大类 ", "金额"]], True), 1
        )
        self.assertIsNone(find_asset_category_column([["客户", "资产类别", "金额"]], True))
        self.assertIsNone(find_asset_category_column([["资产大类", "金额"]], False))


class ExcelTests(unittest.TestCase):
    def test_generated_workbook_content_and_formatting(self) -> None:
        parsed = parse_pasted_table(SAMPLE)
        workbook = load_workbook(build_excel(parsed.rows, True))
        sheet = workbook["整理后数据"]

        self.assertEqual(sheet.freeze_panes, "A2")
        self.assertEqual(sheet.auto_filter.ref, f"A1:D{sheet.max_row}")
        self.assertEqual(sheet["A1"].value, "姓名")
        self.assertTrue(sheet["A1"].font.bold)
        self.assertEqual(sheet["A1"].font.name, "等线")
        self.assertEqual(sheet["A1"].fill.fgColor.rgb, "00D9EAF7")
        self.assertEqual(sheet["A1"].alignment.horizontal, "center")
        self.assertEqual(sheet["C2"].value, 98)
        self.assertEqual(sheet["C2"].alignment.horizontal, "center")
        self.assertEqual(sheet["B2"].value, "00123")
        self.assertEqual(sheet["B2"].data_type, "s")
        self.assertEqual(sheet["B2"].alignment.horizontal, "center")
        self.assertFalse(sheet["B2"].font.bold)
        self.assertEqual(sheet["C3"].value, 0)
        # 三个星号逐字符替换为 000，并按前导零规则保留为文本。
        self.assertEqual(sheet["C4"].value, "000")
        self.assertIsNone(sheet["A5"].value)
        self.assertGreaterEqual(sheet.column_dimensions["A"].width, 8)
        self.assertLessEqual(sheet.column_dimensions["D"].width, 35)
        self.assertTrue(sheet["D2"].alignment.wrap_text)
        self.assertGreater(sheet.row_dimensions[2].height, 20)
        self.assertEqual(sheet["A1"].border.left.style, "medium")
        self.assertEqual(sheet["B2"].border.left.style, "thin")

    def test_asset_categories_use_distinct_fresh_row_colors(self) -> None:
        rows = [
            ["客户", "资产大类", "产品名称", "金额", "占比"],
            ["张三", " 固定收益类 ", "稳健债券", "100", "40%"],
            ["张三", "权益类", "股票基金", "200", "60%"],
            ["现金类", "货币基金", "50"],
            ["另类投资", "黄金", "30"],
            ["保障类", "年金保险", "20"],
            ["合计", "合计", "待分类资产", "300", "100%"],
            ["李四", "固定收益", "近似词不应填色", "10", "10%"],
        ]
        sheet = load_workbook(build_excel(rows, True)).active

        # 整条记录只在实际数据区域 A:E 内填色。
        for cell in sheet[2][0:5]:
            self.assertEqual(cell.fill.fgColor.rgb, "00E8F3EC")
        for cell in sheet[3][0:5]:
            self.assertEqual(cell.fill.fgColor.rgb, "00DDEBF7")
        self.assertIsNone(sheet["F2"].value)
        self.assertEqual(sheet["F2"].fill.fill_type, None)

        # 现金类、另类投资、保障类、合计及近似词均保持原来无填色状态。
        for row_number in (4, 5, 6, 7, 8):
            for cell in sheet[row_number][0:5]:
                self.assertEqual(cell.fill.fill_type, None)

        # 所有数据区域单元格，无论文字、数字或百分比文本，都水平/垂直居中。
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=5):
            for cell in row:
                self.assertEqual(cell.alignment.horizontal, "center")
                self.assertEqual(cell.alignment.vertical, "center")

        # 分类填色不得改变既有字体、字体颜色、边框、尺寸或数字格式。
        for coordinate in ("A2", "B2", "C2", "D2", "E2", "A3", "E3"):
            cell = sheet[coordinate]
            self.assertEqual(cell.font.name, "等线")
            self.assertEqual(cell.font.sz, 10)
            self.assertFalse(cell.font.bold)
            self.assertEqual(cell.font.color, None)
            self.assertEqual(cell.number_format, "@" if isinstance(cell.value, str) else "General")
        self.assertEqual(sheet["B2"]._style.borderId, sheet["B3"]._style.borderId)
        self.assertEqual(sheet["C2"]._style.borderId, sheet["C3"]._style.borderId)
        self.assertEqual(sheet.row_dimensions[2].height, sheet.row_dimensions[3].height)
        for column in ("A", "B", "C", "D", "E"):
            self.assertGreaterEqual(sheet.column_dimensions[column].width, 8)

        # 与不触发分类填色的同结构基准表相比，字体/颜色、边框、尺寸和格式完全一致。
        baseline_rows = [list(row) for row in rows]
        baseline_rows[1][1] = "未识别类别甲"
        baseline_rows[2][1] = "未识别类别乙"
        baseline = load_workbook(build_excel(baseline_rows, True)).active
        for row_number in (2, 3):
            for column_number in range(1, 6):
                styled = sheet.cell(row_number, column_number)
                original = baseline.cell(row_number, column_number)
                self.assertEqual(styled._style.fontId, original._style.fontId)
                self.assertEqual(styled._style.borderId, original._style.borderId)
                self.assertEqual(styled._style.numFmtId, original._style.numFmtId)
                self.assertEqual(styled.font.color, original.font.color)
        self.assertEqual(
            [sheet.column_dimensions[column].width for column in ("A", "B", "C", "D", "E")],
            [baseline.column_dimensions[column].width for column in ("A", "B", "C", "D", "E")],
        )
        self.assertEqual(sheet.row_dimensions[2].height, baseline.row_dimensions[2].height)
        self.assertEqual(sheet.row_dimensions[3].height, baseline.row_dimensions[3].height)

    def test_missing_asset_category_column_still_exports_without_category_fill(self) -> None:
        rows = [
            ["资产类别", "产品名称", "金额"],
            ["固定收益类", "稳健债券", "100"],
            ["权益类", "股票基金", "200"],
        ]
        sheet = load_workbook(build_excel(rows, True)).active
        self.assertIsNone(find_asset_category_column(rows, True))
        for row in sheet.iter_rows(min_row=2, max_row=3, min_col=1, max_col=3):
            for cell in row:
                self.assertEqual(cell.fill.fill_type, None)

    def test_no_header_adds_generic_header(self) -> None:
        rows = [["张三", "001"], ["李四", "2"]]
        sheet = load_workbook(build_excel(rows, False)).active
        self.assertEqual([sheet.cell(1, column).value for column in (1, 2)], ["列1", "列2"])
        self.assertEqual([sheet.cell(2, column).value for column in (1, 2)], ["张三", "001"])
        self.assertEqual(sheet.auto_filter.ref, "A1:B3")

    def test_formula_like_text_is_not_executed(self) -> None:
        sheet = load_workbook(build_excel([["内容"], ["=1+1"]], True), data_only=False).active
        self.assertEqual(sheet["A2"].value, "=1+1")
        self.assertEqual(sheet["A2"].data_type, "s")

    def test_empty_table_raises_friendly_value_error(self) -> None:
        with self.assertRaisesRegex(ValueError, "没有可导出"):
            build_excel([], True)


if __name__ == "__main__":
    unittest.main()
